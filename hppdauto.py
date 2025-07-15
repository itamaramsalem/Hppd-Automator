def run_hppd_comparison_for_date(templates_folder, reports_folder, target_date, output_path):
    from openpyxl import Workbook
    import pandas as pd
    import openpyxl
    import xlrd
    import os
    import re
    from datetime import datetime
    from difflib import get_close_matches
    from openpyxl.styles import PatternFill, Font, Alignment, numbers
    from openpyxl.utils import get_column_letter
    from openpyxl.utils.dataframe import dataframe_to_rows

    def normalize_name(name):
        name = name.lower()
        name = re.sub(r"[^a-z0-9\s]", "", name)
        name = re.sub(r"\s+", " ", name).strip()
        return name

    def extract_core_from_report(report_name):
        if report_name.lower().startswith("total nursing wrkd - "):
            return normalize_name(report_name[21:])
        return normalize_name(report_name)

    def build_template_name_map(template_entries):
        return {entry["cleaned_name"]: entry["facility"] for entry in template_entries}

    def match_report_to_template(report_name, template_name_map, cutoff=0.4):
        core_name = extract_core_from_report(report_name)
        match = get_close_matches(core_name, list(template_name_map.keys()), n=1, cutoff=cutoff)
        return template_name_map[match[0]] if match else None

    # === Parse Templates ===
    template_entries = []
    for filename in os.listdir(templates_folder):
        if not filename.endswith(".xlsx"):
            continue
        filepath = os.path.join(templates_folder, filename)
        wb = openpyxl.load_workbook(filepath, data_only=True)
        try:
            facility_full = wb["1"]["D3"].value
            cleaned_facility = normalize_name(facility_full)
        except:
            continue

        for ws in wb.worksheets:
            try:
                date_cell = ws["B11"].value
                sheet_date = (
                    date_cell.date() if isinstance(date_cell, datetime)
                    else pd.to_datetime(date_cell).date()
                )
                if target_date and sheet_date != datetime.strptime(target_date, "%Y-%m-%d").date():
                    continue
                census = float(ws["E27"].value)
                template_entries.append({
                    "facility": facility_full,
                    "cleaned_name": cleaned_facility,
                    "date": sheet_date,
                    "census": census,
                    "file": filename,
                    "sheet": ws.title
                })
            except:
                continue

    # === Parse Reports ===
    results = []
    template_map = build_template_name_map(template_entries)

    for filename in os.listdir(reports_folder):
        if not filename.endswith(".xls"):
            continue
        filepath = os.path.join(reports_folder, filename)
        try:
            wb = xlrd.open_workbook(filepath)
            ws = wb.sheet_by_name("Sheet3")
            raw_date = ws.cell_value(3, 1)
            report_date = datetime(*xlrd.xldate_as_tuple(raw_date, wb.datemode)).date() if isinstance(raw_date, float) else pd.to_datetime(raw_date).date()
            if target_date and report_date != datetime.strptime(target_date, "%Y-%m-%d").date():
                continue
            report_facility = ws.cell_value(4, 1)
            actual_hours = float(ws.cell_value(13, 7))
            actual_cna_hours = float(ws.cell_value(12, 7))
            actual_rn_lpn_hours = float(ws.cell_value(11, 7)) + float(ws.cell_value(10, 7))

            matched_template_name = match_report_to_template(report_facility, template_map)
            if not matched_template_name:
                continue

            candidates = [
                entry for entry in template_entries
                if entry["facility"] == matched_template_name and entry["date"] == report_date
            ]
            if not candidates:
                continue

            t = candidates[0]
            actual_hppd = actual_hours / t["census"] if t["census"] else None
            actual_cna_hppd = actual_cna_hours / t["census"] if t["census"] else None
            actual_rn_lpn_hppd = actual_rn_lpn_hours / t["census"] if t["census"] else None

            if actual_hppd is None:
                hppd_flag = "Missing Data"
            elif abs(actual_hppd - 3.2) < 0.01:
                hppd_flag = "On Target"
            elif actual_hppd > 3.2:
                hppd_flag = "Over Budget"
            else:
                hppd_flag = "Under Budget"

            results.append({
                "Date": report_date,
                "Report Facility": report_facility,
                "Template Facility": matched_template_name,
                "Template File": t["file"],
                "Sheet": t["sheet"],
                "Actual Hours": actual_hours,
                "Census": t["census"],
                "Actual HPPD": actual_hppd,
                "Actual CNA HPPD": actual_cna_hppd,
                "Actual RN+LPN HPPD": actual_rn_lpn_hppd,
                "HPPD Budget Status": hppd_flag
            })
        except:
            continue

    # === Categorize ===
    df_results = pd.DataFrame(results).round(2)

    df_pool = df_results.copy()
    good_hppd_mask = df_pool["Actual HPPD"].between(3.00, 3.30)
    good_cna_mask = df_pool["Actual CNA HPPD"].between(2.00, 2.06)
    good_rn_mask = df_pool["Actual RN+LPN HPPD"] <= 1.20
    bad_cna_mask = df_pool["Actual CNA HPPD"] < 2.00
    bad_rn_mask = df_pool["Actual RN+LPN HPPD"] > 1.20
    bad_split_mask = bad_cna_mask | bad_rn_mask
    bad_hppd_mask = ~good_hppd_mask

    group1 = df_pool[good_hppd_mask & good_cna_mask & good_rn_mask].copy()
    df_pool = df_pool.drop(index=group1.index)
    group2 = df_pool[good_hppd_mask & bad_split_mask].copy()
    df_pool = df_pool.drop(index=group2.index)
    group3 = df_pool[bad_hppd_mask & bad_split_mask].copy()

    # === Output Excel File ===
    wb = Workbook()
    ws = wb.active
    ws.title = "Categorized Facilities"

    current_row = 1
    section_definitions = [
        ("Good HPPD & Good Split (3.0<HPPD<3.3, 2.00<CNA<2.06, RN+LPN<=1.20)", group1),
        ("Good HPPD & Bad Split (3.0<HPPD<3.3, CNA<2.00, RN+LPN>1.20)", group2),
        ("Bad HPPD & Bad Split (HPPD>3.3 | HPPD<3.0, CNA<2.00, RN+LPN>1.20)", group3),
    ]

    for section_title, df in section_definitions:
        ws.cell(row=current_row, column=1, value=section_title)
        ws.cell(row=current_row, column=1).font = Font(bold=True, size=12)
        current_row += 1

        if df.empty:
            fill = PatternFill("solid", fgColor="F2F2F2") if current_row % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
            cell = ws.cell(row=current_row, column=1, value="No facilities in this category")
            cell.fill = fill
            current_row += 2
            continue

        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, val in enumerate(row, 1):
                cell = ws.cell(row=current_row, column=c_idx, value=val)
                fill = PatternFill("solid", fgColor="F2F2F2") if current_row % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
                cell.fill = fill

                if r_idx == 1:
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                if r_idx > 1 and c_idx == 1 and isinstance(val, datetime):
                    cell.number_format = numbers.FORMAT_DATE_YYYYMMDD2

            current_row += 1
        current_row += 1

    for col in ws.columns:
        max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 50)

    wb.save(output_path)
